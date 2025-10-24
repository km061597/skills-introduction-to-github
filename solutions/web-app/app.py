#!/usr/bin/env python3
"""
Smart Excel to PDF Converter - Web Application
==============================================
Flask-based web application for converting Excel to PDF with intelligent analysis.

Features:
- Browser-based interface (no installation needed)
- Drag-and-drop file upload
- Real-time analysis
- Interactive customization
- Download or preview PDF
- Mobile-friendly responsive design
"""

from flask import Flask, render_template, request, send_file, jsonify, session
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.utils import get_column_letter
import os
import uuid
from pathlib import Path
from datetime import datetime, timedelta
import json
from typing import List, Dict
import tempfile
import shutil

# PDF generation imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape, portrait
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()
app.config['PDF_FOLDER'] = tempfile.mkdtemp()
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'xls'}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_excel(file_path: str) -> List[Dict]:
    """Analyze Excel file and return sheet information"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        analyses = []

        for sheet in wb.worksheets:
            analysis = analyze_sheet(sheet)
            analyses.append(analysis)

        wb.close()
        return analyses

    except Exception as e:
        raise Exception(f"Error analyzing workbook: {str(e)}")


def analyze_sheet(sheet) -> Dict:
    """Analyze a single sheet"""
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Count non-empty cells
    filled_cells = 0
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                filled_cells += 1

    # Calculate data density
    total_cells = max_row * max_col
    data_density = filled_cells / total_cells if total_cells > 0 else 0

    # Detect headers
    has_headers = detect_headers(sheet)

    # Estimate pages
    estimated_pages = max(1, (max_row + 44) // 45)

    # Recommend orientation
    recommended_orientation = "landscape" if max_col > 10 else "portrait"

    # Check for wide columns
    has_wide_columns = max_col > 15

    # Determine scaling
    requires_scaling = has_wide_columns or max_col > 12

    # Calculate optimal scale
    if max_col <= 10:
        optimal_scale = 100
    elif max_col <= 15:
        optimal_scale = 85
    elif max_col <= 20:
        optimal_scale = 75
    else:
        optimal_scale = 65

    return {
        'sheet_name': sheet.title,
        'row_count': max_row,
        'column_count': max_col,
        'used_range': f"A1:{get_column_letter(max_col)}{max_row}",
        'has_headers': has_headers,
        'data_density': round(data_density * 100, 1),
        'estimated_pages': estimated_pages,
        'recommended_orientation': recommended_orientation,
        'has_wide_columns': has_wide_columns,
        'requires_scaling': requires_scaling,
        'optimal_scale': optimal_scale
    }


def detect_headers(sheet) -> bool:
    """Detect if first row contains headers"""
    if sheet.max_row < 2:
        return False

    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]

    bold_count = 0
    text_count = 0

    for cell in first_row:
        if cell.value is not None:
            text_count += 1
            if cell.font and cell.font.bold:
                bold_count += 1

    return (bold_count > text_count * 0.5) or (text_count > len(first_row) * 0.7)


def generate_pdf(excel_path: str, output_path: str, settings: Dict) -> bool:
    """Generate PDF from Excel with specified settings"""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Get orientation
        if settings.get('orientation', 'auto').lower() == 'landscape':
            page_size = landscape(letter)
        else:
            page_size = portrait(letter)

        # Get margins
        margins = settings.get('margins', {})
        left_margin = margins.get('left', 0.5) * inch
        right_margin = margins.get('right', 0.5) * inch
        top_margin = margins.get('top', 0.75) * inch
        bottom_margin = margins.get('bottom', 0.75) * inch

        # Create PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=page_size,
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=top_margin,
            bottomMargin=bottom_margin,
            title=Path(excel_path).stem
        )

        elements = []
        styles = getSampleStyleSheet()

        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=20,
            alignment=1
        )

        selected_sheets = settings.get('selected_sheets', [])
        include_gridlines = settings.get('include_gridlines', False)

        for idx, sheet_name in enumerate(selected_sheets):
            sheet = wb[sheet_name]

            # Add sheet title
            title = Paragraph(f"<b>{sheet_name}</b>", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2 * inch))

            # Convert sheet to table data
            data = []
            for row in sheet.iter_rows(values_only=True):
                processed_row = [
                    str(cell) if cell is not None else ""
                    for cell in row
                ]
                data.append(processed_row)

            if not data:
                elements.append(Paragraph("<i>Empty sheet</i>", styles['Normal']))
                elements.append(PageBreak())
                continue

            # Create table
            table = Table(data)

            # Apply styling
            table_style = [
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]

            if include_gridlines:
                table_style.extend([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ])
            else:
                table_style.extend([
                    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1f4788')),
                ])

            table.setStyle(TableStyle(table_style))
            elements.append(table)

            # Add page break between sheets
            if idx < len(selected_sheets) - 1:
                elements.append(PageBreak())

        # Build PDF
        if settings.get('include_headers', True):
            doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        else:
            doc.build(elements)

        wb.close()
        return True

    except Exception as e:
        raise Exception(f"Error generating PDF: {str(e)}")


def add_page_number(canvas, doc):
    """Add page numbers and footer to PDF"""
    page_num = canvas.getPageNumber()
    text = f"Page {page_num}"
    canvas.saveState()
    canvas.setFont('Helvetica', 9)
    canvas.drawRightString(doc.pagesize[0] - 0.5 * inch, 0.5 * inch, text)
    canvas.drawString(0.5 * inch, 0.5 * inch, datetime.now().strftime("%Y-%m-%d %H:%M"))
    canvas.restoreState()


@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and analysis"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xlsm, or .xls'}), 400

    try:
        # Save file
        file_id = str(uuid.uuid4())
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")
        file.save(file_path)

        # Analyze file
        analyses = analyze_excel(file_path)

        # Store in session
        session['file_id'] = file_id
        session['filename'] = filename
        session['file_path'] = file_path
        session['analyses'] = analyses

        return jsonify({
            'success': True,
            'file_id': file_id,
            'filename': filename,
            'analyses': analyses
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate', methods=['POST'])
def generate():
    """Generate PDF with specified settings"""
    if 'file_path' not in session:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        settings = request.json

        # Generate PDF
        pdf_id = str(uuid.uuid4())
        pdf_filename = f"{Path(session['filename']).stem}_{pdf_id}.pdf"
        pdf_path = os.path.join(app.config['PDF_FOLDER'], pdf_filename)

        generate_pdf(session['file_path'], pdf_path, settings)

        # Store PDF info in session
        session['pdf_path'] = pdf_path
        session['pdf_filename'] = pdf_filename

        return jsonify({
            'success': True,
            'pdf_id': pdf_id,
            'filename': pdf_filename
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download/<pdf_id>')
def download_pdf(pdf_id):
    """Download generated PDF"""
    if 'pdf_path' not in session:
        return "PDF not found", 404

    try:
        return send_file(
            session['pdf_path'],
            as_attachment=True,
            download_name=f"{Path(session['filename']).stem}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        return str(e), 500


@app.route('/preview/<pdf_id>')
def preview_pdf(pdf_id):
    """Preview PDF in browser"""
    if 'pdf_path' not in session:
        return "PDF not found", 404

    try:
        return send_file(
            session['pdf_path'],
            mimetype='application/pdf'
        )
    except Exception as e:
        return str(e), 500


@app.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'pdf_available': PDF_AVAILABLE
    })


def cleanup_old_files():
    """Cleanup old temporary files"""
    # In production, implement a proper cleanup mechanism
    pass


if __name__ == '__main__':
    print("=" * 60)
    print("Smart Excel to PDF Converter - Web Application")
    print("=" * 60)
    print(f"\nServer starting...")
    print(f"Access the application at: http://localhost:5000")
    print(f"\nPress Ctrl+C to stop the server")
    print("=" * 60)

    app.run(debug=True, host='0.0.0.0', port=5000)
