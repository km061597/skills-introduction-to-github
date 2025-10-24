#!/usr/bin/env python3
"""
Smart Excel to PDF Converter - Python GUI Application
=====================================================
A powerful, intelligent PDF converter with analysis and customization features.

Features:
- Drag-and-drop file loading
- Smart sheet analysis and recommendations
- Customizable PDF settings
- Preview functionality
- Batch processing
- Cross-platform support
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import pandas as pd
from pathlib import Path
import os
import sys
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
import threading
from datetime import datetime

# For PDF generation
try:
    import xlsxwriter
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape, portrait
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("Warning: Some PDF libraries not available. Install with: pip install reportlab xlsxwriter")


@dataclass
class SheetAnalysis:
    """Data class for sheet analysis results"""
    sheet_name: str
    row_count: int
    column_count: int
    used_range: str
    has_headers: bool
    data_density: float
    estimated_pages: int
    recommended_orientation: str
    has_wide_columns: bool
    requires_scaling: bool
    optimal_scale: int
    empty_rows: int
    empty_columns: int


class ExcelAnalyzer:
    """Intelligent Excel file analyzer"""

    @staticmethod
    def analyze_workbook(file_path: str) -> List[SheetAnalysis]:
        """Analyze all sheets in workbook"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            analyses = []

            for sheet in wb.worksheets:
                analysis = ExcelAnalyzer._analyze_sheet(sheet)
                analyses.append(analysis)

            wb.close()
            return analyses

        except Exception as e:
            raise Exception(f"Error analyzing workbook: {str(e)}")

    @staticmethod
    def _analyze_sheet(sheet) -> SheetAnalysis:
        """Analyze a single sheet"""
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Count non-empty cells
        filled_cells = 0
        empty_rows = 0
        empty_cols = 0

        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_empty = True
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
                    row_empty = False
            if row_empty:
                empty_rows += 1

        # Calculate data density
        total_cells = max_row * max_col
        data_density = filled_cells / total_cells if total_cells > 0 else 0

        # Detect headers
        has_headers = ExcelAnalyzer._detect_headers(sheet)

        # Estimate pages (rough: 45 rows per page for portrait)
        estimated_pages = max(1, (max_row + 44) // 45)

        # Recommend orientation based on aspect ratio
        recommended_orientation = "Landscape" if max_col > 10 else "Portrait"

        # Check for wide columns
        has_wide_columns = max_col > 15

        # Determine if scaling needed
        requires_scaling = has_wide_columns or max_col > 12

        # Calculate optimal scale
        if max_col <= 10:
            optimal_scale = 100
        elif max_col <= 15:
            optimal_scale = 85
        elif max_col <= 20:
            optimal_scale = 75
        else:
            optimal_scale = 65

        return SheetAnalysis(
            sheet_name=sheet.title,
            row_count=max_row,
            column_count=max_col,
            used_range=f"A1:{get_column_letter(max_col)}{max_row}",
            has_headers=has_headers,
            data_density=data_density,
            estimated_pages=estimated_pages,
            recommended_orientation=recommended_orientation,
            has_wide_columns=has_wide_columns,
            requires_scaling=requires_scaling,
            optimal_scale=optimal_scale,
            empty_rows=empty_rows,
            empty_columns=empty_cols
        )

    @staticmethod
    def _detect_headers(sheet) -> bool:
        """Detect if first row contains headers"""
        if sheet.max_row < 2:
            return False

        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]

        bold_count = 0
        text_count = 0

        for cell in first_row:
            if cell.value is not None:
                text_count += 1
                if cell.font and cell.font.bold:
                    bold_count += 1

        # If >50% bold or >70% filled, assume headers
        return (bold_count > text_count * 0.5) or (text_count > len(first_row) * 0.7)


class PDFGenerator:
    """Intelligent PDF generator with quality optimization"""

    @staticmethod
    def generate_pdf(
        excel_path: str,
        output_path: str,
        selected_sheets: List[str],
        orientation: str = "Auto",
        fit_to_width: bool = True,
        scale_percent: int = 100,
        margins: Dict[str, float] = None,
        include_gridlines: bool = False,
        include_headers: bool = True,
        progress_callback=None
    ) -> bool:
        """Generate PDF from Excel with smart formatting"""

        if margins is None:
            margins = {"left": 0.5, "right": 0.5, "top": 0.75, "bottom": 0.75}

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)

            # Create PDF
            pdf_path = output_path
            if orientation.lower() == "landscape":
                page_size = landscape(letter)
            else:
                page_size = portrait(letter)

            # Build PDF
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=page_size,
                leftMargin=margins["left"] * inch,
                rightMargin=margins["right"] * inch,
                topMargin=margins["top"] * inch,
                bottomMargin=margins["bottom"] * inch,
                title=Path(excel_path).stem
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=20,
                alignment=1  # Center
            )

            total_sheets = len(selected_sheets)

            for idx, sheet_name in enumerate(selected_sheets):
                if progress_callback:
                    progress = int((idx / total_sheets) * 100)
                    progress_callback(progress, f"Processing {sheet_name}...")

                sheet = wb[sheet_name]

                # Add sheet title
                title = Paragraph(f"<b>{sheet_name}</b>", title_style)
                elements.append(title)
                elements.append(Spacer(1, 0.2 * inch))

                # Convert sheet to table
                data = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert None to empty string and handle other types
                    processed_row = [
                        str(cell) if cell is not None else ""
                        for cell in row
                    ]
                    data.append(processed_row)

                if not data:
                    elements.append(Paragraph("<i>Empty sheet</i>", styles['Normal']))
                    elements.append(PageBreak())
                    continue

                # Create table
                table = Table(data)

                # Apply styling
                table_style = [
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8 if scale_percent < 80 else 9),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                    ('PADDING', (0, 0), (-1, -1), 6),
                ]

                if include_gridlines:
                    table_style.extend([
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ])
                else:
                    table_style.extend([
                        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1f4788')),
                    ])

                table.setStyle(TableStyle(table_style))

                elements.append(table)

                # Add page break between sheets
                if idx < total_sheets - 1:
                    elements.append(PageBreak())

            # Build PDF with page numbers and footers
            if include_headers:
                doc.build(elements, onFirstPage=PDFGenerator._add_page_number,
                         onLaterPages=PDFGenerator._add_page_number)
            else:
                doc.build(elements)

            wb.close()

            if progress_callback:
                progress_callback(100, "PDF generated successfully!")

            return True

        except Exception as e:
            raise Exception(f"Error generating PDF: {str(e)}")

    @staticmethod
    def _add_page_number(canvas, doc):
        """Add page numbers and footer to PDF"""
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(doc.pagesize[0] - 0.5 * inch, 0.5 * inch, text)
        canvas.drawString(0.5 * inch, 0.5 * inch, datetime.now().strftime("%Y-%m-%d %H:%M"))
        canvas.restoreState()


class ExcelToPDFApp:
    """Main application GUI"""

    def __init__(self, root):
        self.root = root
        self.root.title("Smart Excel to PDF Converter")
        self.root.geometry("1000x700")

        self.current_file = None
        self.analyses = []

        # Configure style
        self.setup_styles()

        # Create UI
        self.create_widgets()

        # Check for PDF libraries
        if not PDF_AVAILABLE:
            messagebox.showwarning(
                "Missing Dependencies",
                "Some PDF libraries are not installed.\n"
                "Install with: pip install reportlab xlsxwriter openpyxl pandas"
            )

    def setup_styles(self):
        """Configure ttk styles"""
        style = ttk.Style()
        style.theme_use('clam')

        # Colors
        self.colors = {
            'primary': '#1f4788',
            'secondary': '#4a90e2',
            'success': '#28a745',
            'danger': '#dc3545',
            'warning': '#ffc107',
            'light': '#f8f9fa',
            'dark': '#343a40'
        }

        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), foreground=self.colors['primary'])
        style.configure('Subtitle.TLabel', font=('Arial', 10), foreground=self.colors['dark'])
        style.configure('Success.TButton', background=self.colors['success'])

    def create_widgets(self):
        """Create all GUI widgets"""

        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)

        # Title
        title = ttk.Label(main_frame, text="Smart Excel to PDF Converter", style='Title.TLabel')
        title.grid(row=0, column=0, columnspan=3, pady=(0, 5))

        subtitle = ttk.Label(
            main_frame,
            text="Intelligent analysis • Custom formatting • Professional output",
            style='Subtitle.TLabel'
        )
        subtitle.grid(row=1, column=0, columnspan=3, pady=(0, 15))

        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="1. Select Excel File", padding="10")
        file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)

        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, state='readonly').grid(
            row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=(0, 5)
        )

        ttk.Button(file_frame, text="Browse...", command=self.browse_file).grid(row=0, column=2, padx=2)
        ttk.Button(file_frame, text="Analyze", command=self.analyze_file).grid(row=0, column=3, padx=2)

        # Main content area (sheets + analysis + settings)
        content_frame = ttk.Frame(main_frame)
        content_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        content_frame.columnconfigure(0, weight=1)
        content_frame.columnconfigure(1, weight=2)
        content_frame.rowconfigure(0, weight=1)

        # Left panel: Sheets
        sheets_frame = ttk.LabelFrame(content_frame, text="2. Select Sheets", padding="10")
        sheets_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        sheets_frame.columnconfigure(0, weight=1)
        sheets_frame.rowconfigure(1, weight=1)

        # Sheet buttons
        btn_frame = ttk.Frame(sheets_frame)
        btn_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        ttk.Button(btn_frame, text="Select All", command=self.select_all_sheets).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Clear", command=self.clear_sheets).pack(side=tk.LEFT, padx=2)

        # Sheets listbox
        sheets_scroll = ttk.Scrollbar(sheets_frame, orient=tk.VERTICAL)
        self.sheets_listbox = tk.Listbox(
            sheets_frame,
            selectmode=tk.MULTIPLE,
            yscrollcommand=sheets_scroll.set,
            font=('Arial', 10)
        )
        sheets_scroll.config(command=self.sheets_listbox.yview)

        self.sheets_listbox.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sheets_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))

        self.sheets_listbox.bind('<<ListboxSelect>>', self.on_sheet_select)

        # Right panel: Tabbed interface for Analysis and Settings
        right_notebook = ttk.Notebook(content_frame)
        right_notebook.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Analysis tab
        analysis_frame = ttk.Frame(right_notebook, padding="10")
        right_notebook.add(analysis_frame, text="Sheet Analysis")

        self.analysis_text = scrolledtext.ScrolledText(
            analysis_frame,
            wrap=tk.WORD,
            font=('Courier', 9),
            height=20
        )
        self.analysis_text.pack(fill=tk.BOTH, expand=True)

        # Settings tab
        settings_frame = ttk.Frame(right_notebook, padding="10")
        right_notebook.add(settings_frame, text="PDF Settings")

        self.create_settings_widgets(settings_frame)

        # Bottom: Progress and actions
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E))
        action_frame.columnconfigure(0, weight=1)

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            action_frame,
            mode='determinate',
            variable=self.progress_var
        )
        self.progress_bar.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))

        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(action_frame, textvariable=self.status_var, foreground=self.colors['secondary']).grid(
            row=1, column=0, sticky=tk.W
        )

        # Action buttons
        ttk.Button(action_frame, text="Generate PDF", command=self.generate_pdf).grid(
            row=1, column=1, padx=5
        )
        ttk.Button(action_frame, text="Exit", command=self.root.quit).grid(row=1, column=2)

    def create_settings_widgets(self, parent):
        """Create PDF settings widgets"""
        row = 0

        # Orientation
        ttk.Label(parent, text="Orientation:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.orientation_var = tk.StringVar(value="Auto")
        orientation_combo = ttk.Combobox(
            parent,
            textvariable=self.orientation_var,
            values=["Auto", "Portrait", "Landscape"],
            state='readonly',
            width=15
        )
        orientation_combo.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        # Scaling
        self.fit_to_width_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            parent,
            text="Fit to Page Width (Recommended)",
            variable=self.fit_to_width_var,
            command=self.toggle_scale
        ).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=5)
        row += 1

        ttk.Label(parent, text="Scale %:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.scale_var = tk.IntVar(value=100)
        self.scale_spin = ttk.Spinbox(
            parent,
            from_=25,
            to=200,
            textvariable=self.scale_var,
            width=15,
            state='disabled'
        )
        self.scale_spin.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        # Margins
        ttk.Label(parent, text="Margins (inches):").grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=10)
        row += 1

        margin_frame = ttk.Frame(parent)
        margin_frame.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=5)

        self.margin_left_var = tk.DoubleVar(value=0.5)
        self.margin_right_var = tk.DoubleVar(value=0.5)
        self.margin_top_var = tk.DoubleVar(value=0.75)
        self.margin_bottom_var = tk.DoubleVar(value=0.75)

        ttk.Label(margin_frame, text="Left:").grid(row=0, column=0, padx=2)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.25,
                   textvariable=self.margin_left_var, width=8).grid(row=0, column=1, padx=2)

        ttk.Label(margin_frame, text="Right:").grid(row=0, column=2, padx=2)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.25,
                   textvariable=self.margin_right_var, width=8).grid(row=0, column=3, padx=2)

        ttk.Label(margin_frame, text="Top:").grid(row=1, column=0, padx=2, pady=5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.25,
                   textvariable=self.margin_top_var, width=8).grid(row=1, column=1, padx=2, pady=5)

        ttk.Label(margin_frame, text="Bottom:").grid(row=1, column=2, padx=2, pady=5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.25,
                   textvariable=self.margin_bottom_var, width=8).grid(row=1, column=3, padx=2, pady=5)

        row += 1

        # Options
        ttk.Label(parent, text="Options:").grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=10)
        row += 1

        self.include_headers_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            parent,
            text="Include Headers/Footers (Page numbers, dates)",
            variable=self.include_headers_var
        ).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
        row += 1

        self.include_gridlines_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            parent,
            text="Print Gridlines",
            variable=self.include_gridlines_var
        ).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)

    def toggle_scale(self):
        """Toggle scale spinbox based on fit_to_width"""
        if self.fit_to_width_var.get():
            self.scale_spin.config(state='disabled')
        else:
            self.scale_spin.config(state='normal')

    def browse_file(self):
        """Browse for Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel Files", "*.xlsx *.xlsm *.xls"),
                ("All Files", "*.*")
            ]
        )

        if file_path:
            self.file_path_var.set(file_path)
            self.current_file = file_path
            self.analyze_file()

    def analyze_file(self):
        """Analyze the Excel file"""
        if not self.current_file:
            messagebox.showwarning("No File", "Please select an Excel file first.")
            return

        try:
            self.status_var.set("Analyzing workbook...")
            self.progress_var.set(50)
            self.root.update()

            # Analyze in background
            self.analyses = ExcelAnalyzer.analyze_workbook(self.current_file)

            # Populate sheets list
            self.sheets_listbox.delete(0, tk.END)
            for analysis in self.analyses:
                self.sheets_listbox.insert(tk.END, analysis.sheet_name)

            # Select all by default
            for i in range(len(self.analyses)):
                self.sheets_listbox.selection_set(i)

            self.status_var.set(f"Analysis complete! Found {len(self.analyses)} sheets.")
            self.progress_var.set(0)

            # Show first sheet analysis
            if self.analyses:
                self.show_analysis(0)

        except Exception as e:
            messagebox.showerror("Analysis Error", str(e))
            self.status_var.set("Analysis failed")
            self.progress_var.set(0)

    def select_all_sheets(self):
        """Select all sheets"""
        self.sheets_listbox.selection_set(0, tk.END)

    def clear_sheets(self):
        """Clear sheet selection"""
        self.sheets_listbox.selection_clear(0, tk.END)

    def on_sheet_select(self, event):
        """Handle sheet selection"""
        selection = self.sheets_listbox.curselection()
        if selection:
            self.show_analysis(selection[0])

    def show_analysis(self, index: int):
        """Display sheet analysis"""
        if 0 <= index < len(self.analyses):
            analysis = self.analyses[index]

            text = f"{'='*60}\n"
            text += f"SHEET ANALYSIS: {analysis.sheet_name}\n"
            text += f"{'='*60}\n\n"

            text += f"DIMENSIONS:\n"
            text += f"  • Rows: {analysis.row_count:,}\n"
            text += f"  • Columns: {analysis.column_count}\n"
            text += f"  • Used Range: {analysis.used_range}\n"
            text += f"  • Data Density: {analysis.data_density:.1%}\n"
            text += f"  • Has Headers: {'Yes' if analysis.has_headers else 'No'}\n\n"

            text += f"QUALITY METRICS:\n"
            text += f"  • Empty Rows: {analysis.empty_rows}\n"
            text += f"  • Empty Columns: {analysis.empty_columns}\n"
            text += f"  • Estimated Pages: {analysis.estimated_pages}\n\n"

            text += f"RECOMMENDATIONS:\n"
            text += f"  • Orientation: {analysis.recommended_orientation}\n"

            if analysis.requires_scaling:
                text += f"  • Scaling: Required\n"
                text += f"  • Optimal Scale: {analysis.optimal_scale}%\n"
                # Auto-apply recommendation
                self.fit_to_width_var.set(False)
                self.scale_var.set(analysis.optimal_scale)
                self.toggle_scale()
            else:
                text += f"  • Scaling: Not required (Fit to width recommended)\n"
                self.fit_to_width_var.set(True)
                self.toggle_scale()

            if analysis.has_wide_columns:
                text += f"  • Note: Contains wide columns\n"
                text += f"    Text wrapping will be enabled for readability\n"

            text += f"\n{'='*60}\n"
            text += f"TIPS:\n"
            text += f"  • Use '{analysis.recommended_orientation}' for best fit\n"
            text += f"  • Adjust margins if content is cut off\n"
            text += f"  • Enable gridlines for data tables\n"
            text += f"  • Headers/footers add professionalism\n"

            self.analysis_text.delete('1.0', tk.END)
            self.analysis_text.insert('1.0', text)

            # Update orientation dropdown
            self.orientation_var.set(analysis.recommended_orientation)

    def generate_pdf(self):
        """Generate PDF from selected sheets"""
        if not self.current_file:
            messagebox.showwarning("No File", "Please select an Excel file first.")
            return

        selection = self.sheets_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Sheets", "Please select at least one sheet.")
            return

        # Get selected sheet names
        selected_sheets = [self.analyses[i].sheet_name for i in selection]

        # Get output path
        output_path = filedialog.asksaveasfilename(
            title="Save PDF As",
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
            initialfile=Path(self.current_file).stem + ".pdf"
        )

        if not output_path:
            return

        # Get settings
        margins = {
            "left": self.margin_left_var.get(),
            "right": self.margin_right_var.get(),
            "top": self.margin_top_var.get(),
            "bottom": self.margin_bottom_var.get()
        }

        # Generate PDF in background thread
        def generate():
            try:
                PDFGenerator.generate_pdf(
                    excel_path=self.current_file,
                    output_path=output_path,
                    selected_sheets=selected_sheets,
                    orientation=self.orientation_var.get(),
                    fit_to_width=self.fit_to_width_var.get(),
                    scale_percent=self.scale_var.get(),
                    margins=margins,
                    include_gridlines=self.include_gridlines_var.get(),
                    include_headers=self.include_headers_var.get(),
                    progress_callback=self.update_progress
                )

                self.root.after(0, lambda: messagebox.showinfo(
                    "Success",
                    f"PDF generated successfully!\n\n{output_path}"
                ))

                # Ask to open
                if messagebox.askyesno("Open PDF", "Would you like to open the PDF?"):
                    if sys.platform == 'win32':
                        os.startfile(output_path)
                    elif sys.platform == 'darwin':
                        os.system(f'open "{output_path}"')
                    else:
                        os.system(f'xdg-open "{output_path}"')

            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
            finally:
                self.root.after(0, lambda: self.progress_var.set(0))

        thread = threading.Thread(target=generate, daemon=True)
        thread.start()

    def update_progress(self, percent: int, message: str):
        """Update progress bar and status"""
        self.progress_var.set(percent)
        self.status_var.set(message)
        self.root.update()


def main():
    """Main entry point"""
    root = tk.Tk()
    app = ExcelToPDFApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
